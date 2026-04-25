#!/usr/bin/env python3
"""One-time secure importer from APAN NIBASH Excel workbook to SQLite DB.

Usage:
  ./.venv/bin/python import_excel_to_db.py --file "APAN NIBASH-21.xlsx" --apply
  ./.venv/bin/python import_excel_to_db.py --file "APAN NIBASH-21.xlsx" --dry-run
Safety:
- Reads local workbook only (no network usage).
- Creates a DB backup before applying.
- Uses a single transaction; rolls back on failure.
"""

from __future__ import annotations

import argparse
import shutil
import sqlite3
import json
from datetime import datetime
from pathlib import Path
from typing import Dict, Tuple

import openpyxl

from database import DB_PATH, get_connection, init_database


EXPENSE_ACCOUNT_MAP: Dict[str, str] = {
    "preliminary": "403",
    "consultant": "209",
    "mixture": "210",
    "vibrator": "210",
    "jahan": "404",
    "sabuj": "404",
    "rafiqul": "404",
    "pilling": "404",
    "piling": "404",
    "rod": "301",
    "steel": "301",
    "cement": "302",
    "stone": "305",
    "sand": "303",
    "carraying": "309",
    "carrying": "309",
    "deep tubwell": "405",
    "gas connection": "406",
    "electricity": "202",
    "land registration": "402",
    "salary": "206",
    "allowances": "206",
    "office rent": "201",
    "entertainment": "205",
    "printing": "204",
    "stationery": "204",
    "conveyance": "208",
    "legal": "209",
    "sanitary": "311",
    "plumbing": "311",
    "electric materials": "312",
    "tiles": "313",
    "window": "314",
    "grill": "314",
    "doors": "315",
    "chawkath": "315",
    "paint": "316",
    "thai glass": "317",
    "commission": "408",
    "loan": "409",
    "refund": "409",
    "holding tax": "411",
    "lift": "412",
}


def normalize_name(name: str) -> str:
    return " ".join(str(name or "").strip().split()).lower()


def map_expense_code(particular: str) -> str:
    key = normalize_name(particular)
    for token, code in EXPENSE_ACCOUNT_MAP.items():
        if token in key:
            return code
    return "210"


def ensure_db_ready() -> None:
    init_database()


def backup_db() -> Path:
    backup_path = DB_PATH.with_name(f"apan_nibash.backup.{datetime.now().strftime('%Y%m%d_%H%M%S')}.db")
    shutil.copy2(DB_PATH, backup_path)
    return backup_path


def archive_workbook_rows(wb, cursor: sqlite3.Cursor) -> int:
    archived = 0
    for ws in wb.worksheets:
        for row_no in range(1, ws.max_row + 1):
            values = [ws.cell(row_no, col).value for col in range(1, ws.max_column + 1)]
            if not any(value not in (None, '', ' ') for value in values):
                continue

            title = None
            for value in values:
                if value not in (None, '', ' '):
                    title = str(value).strip()
                    break

            date_value = None
            amount = 0.0
            amount_2 = 0.0
            for value in values:
                if isinstance(value, datetime) and date_value is None:
                    date_value = value.strftime('%Y-%m-%d')
                elif isinstance(value, (int, float)):
                    if amount == 0.0:
                        amount = float(value)
                    elif amount_2 == 0.0:
                        amount_2 = float(value)

            cursor.execute(
                """
                INSERT OR IGNORE INTO source_rows
                (sheet_name, row_no, record_type, title, date_value, amount, amount_2, row_json)
                VALUES (?, ?, 'raw', ?, ?, ?, ?, ?)
                """,
                (
                    ws.title,
                    row_no,
                    title,
                    date_value,
                    amount,
                    amount_2,
                    json.dumps([str(v) if isinstance(v, datetime) else v for v in values], default=str),
                ),
            )
            archived += 1

    return archived


def next_import_voucher_no(cursor: sqlite3.Cursor, voucher_type: str) -> str:
    cursor.execute(
        "SELECT COUNT(*) FROM vouchers WHERE voucher_no LIKE ?",
        (f"{voucher_type}-IMP-%",),
    )
    count = int(cursor.fetchone()[0])
    return f"{voucher_type}-IMP-{count + 1:06d}"


def insert_voucher(
    cursor: sqlite3.Cursor,
    *,
    voucher_type: str,
    account_code: str,
    description: str,
    amount: float,
    voucher_date: str,
    notes: str,
    payee_payor: str = "",
) -> None:
    if amount <= 0:
        return

    voucher_no = next_import_voucher_no(cursor, voucher_type)
    debit_amount = amount if voucher_type == "PV" else 0.0
    credit_amount = amount if voucher_type == "RV" else 0.0

    cursor.execute(
        """
        INSERT INTO vouchers
        (voucher_no, date, voucher_type, account_code, description,
         debit_amount, credit_amount, reference_no, payee_payor, notes)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            voucher_no,
            voucher_date,
            voucher_type,
            account_code,
            description[:255],
            debit_amount,
            credit_amount,
            "",
            payee_payor[:255],
            notes[:255],
        ),
    )


def as_amount(value) -> int:
    """Convert amount to INTEGER cents for storage"""
    if value is None:
        return 0
    try:
        return int(round(float(value) * 100))
    except Exception:
        return 0


def date_from_cell(cell_value, fallback: str) -> str:
    if isinstance(cell_value, datetime):
        return cell_value.strftime("%Y-%m-%d")
    return fallback


def import_expense_section(ws, cursor: sqlite3.Cursor) -> int:
    imported = 0
    # Row 5-77 contains consolidated expenditure lines.
    for row in range(5, 78):
        particular = ws.cell(row, 2).value
        if not particular:
            continue

        name = str(particular).strip()
        if name.lower().startswith("total"):
            continue

        opening = as_amount(ws.cell(row, 3).value)
        if opening > 0:
            insert_voucher(
                cursor,
                voucher_type="PV",
                account_code=map_expense_code(name),
                description=name,
                amount=opening,
                voucher_date="2021-12-31",
                notes="Imported opening cumulative up to Dec-2021",
            )
            imported += 1

        for col in range(4, 10):
            amount = as_amount(ws.cell(row, col).value)
            if amount <= 0:
                continue
            month_cell = ws.cell(4, col).value
            insert_voucher(
                cursor,
                voucher_type="PV",
                account_code=map_expense_code(name),
                description=name,
                amount=amount,
                voucher_date=date_from_cell(month_cell, "2022-01-01"),
                notes="Imported monthly amount from Year-2022(1)",
            )
            imported += 1

    return imported


def import_investments(ws, cursor: sqlite3.Cursor) -> Tuple[int, int]:
    inserted_investments = 0
    inserted_vouchers = 0

    # Received section: 84-90
    for row in range(84, 91):
        name = ws.cell(row, 2).value
        total = as_amount(ws.cell(row, 10).value)
        if not name or total <= 0:
            continue
        nm = str(name).strip()
        if nm.lower().startswith("total"):
            continue

        cursor.execute(
            """
            INSERT INTO investments (name, type, amount, date, status)
            VALUES (?, 'RECEIVED', ?, ?, 'ACTIVE')
            """,
            (nm, total, "2022-06-30"),
        )
        inserted_investments += 1

        insert_voucher(
            cursor,
            voucher_type="RV",
            account_code="102",
            description=f"Investment received: {nm}",
            amount=total,
            voucher_date="2022-06-30",
            notes="Imported investment received summary",
            payee_payor=nm,
        )
        inserted_vouchers += 1

    # Payments section: 100-106
    for row in range(100, 107):
        name = ws.cell(row, 2).value
        total = as_amount(ws.cell(row, 10).value)
        if not name or total <= 0:
            continue
        nm = str(name).strip()
        if nm.lower().startswith("total"):
            continue

        cursor.execute(
            """
            INSERT INTO investments (name, type, amount, date, status)
            VALUES (?, 'PAID', ?, ?, 'ACTIVE')
            """,
            (nm, total, "2022-06-30"),
        )
        inserted_investments += 1

        insert_voucher(
            cursor,
            voucher_type="PV",
            account_code="409",
            description=f"Investment repayment: {nm}",
            amount=total,
            voucher_date="2022-06-30",
            notes="Imported investment payment summary",
            payee_payor=nm,
        )
        inserted_vouchers += 1

    return inserted_investments, inserted_vouchers


def import_flatholders(ws, cursor: sqlite3.Cursor) -> Tuple[int, int]:
    inserted_holders = 0
    inserted_payments = 0

    # Consolidated flatholder section: 155-205
    for row in range(155, 206):
        serial = ws.cell(row, 1).value
        name = ws.cell(row, 2).value
        total_paid = as_amount(ws.cell(row, 10).value)

        if not serial or not name:
            continue
        if str(name).strip().lower().startswith("total"):
            continue
        if total_paid <= 0:
            continue

        nm = str(name).strip()
        flat_unit = f"Flat-{int(serial)}"

        cursor.execute(
            """
            INSERT OR IGNORE INTO flatholders
            (serial_no, name, phone, email, address, flat_unit, total_amount, paid_amount)
            VALUES (?, ?, '', '', '', ?, ?, ?)
            """,
            (int(serial), nm, flat_unit, total_paid, total_paid),
        )

        # Ensure paid amount is at least imported value.
        cursor.execute(
            """
            UPDATE flatholders
            SET paid_amount = CASE WHEN paid_amount < ? THEN ? ELSE paid_amount END,
                total_amount = CASE WHEN total_amount < ? THEN ? ELSE total_amount END
            WHERE serial_no = ?
            """,
            (total_paid, total_paid, total_paid, total_paid, int(serial)),
        )

        cursor.execute("SELECT id FROM flatholders WHERE serial_no = ?", (int(serial),))
        holder = cursor.fetchone()
        if holder:
            cursor.execute(
                """
                INSERT INTO flatholder_payments
                (flatholder_id, payment_date, amount, payment_type, notes)
                VALUES (?, ?, ?, 'INSTALLMENT', ?)
                """,
                (
                    int(holder[0]),
                    "2022-06-30",
                    total_paid,
                    "Imported cumulative paid amount up to 30/06/2022",
                ),
            )
            inserted_payments += 1

        inserted_holders += 1

    return inserted_holders, inserted_payments


def import_income_statement(ws, cursor: sqlite3.Cursor) -> int:
    inserted = 0

    # Cumulative income row
    cumulative_row = 267
    bank = as_amount(ws.cell(cumulative_row, 3).value)
    fdr = as_amount(ws.cell(cumulative_row, 4).value)
    sale = as_amount(ws.cell(cumulative_row, 5).value)
    charges = as_amount(ws.cell(cumulative_row, 6).value)
    rent = as_amount(ws.cell(cumulative_row, 7).value)

    cumulative_items = [
        ("103", "Bank Profit (cumulative)", bank),
        ("103", "FDR Profit (cumulative)", fdr),
        ("104", "Sale of scrap/wastage (cumulative)", sale),
        ("105", "Service charges (cumulative)", charges),
        ("105", "Rent income (cumulative)", rent),
    ]

    for code, desc, amount in cumulative_items:
        if amount > 0:
            insert_voucher(
                cursor,
                voucher_type="RV",
                account_code=code,
                description=desc,
                amount=amount,
                voucher_date="2021-12-31",
                notes="Imported cumulative income up to Dec-2021",
            )
            inserted += 1

    # Monthly rows 268-273 for Jan-Jun 2022
    for row in range(268, 274):
        month_label = ws.cell(row, 2).value
        if not month_label:
            continue
        date_guess = f"2022-{row-267:02d}-01"
        vals = [
            ("103", "Bank Profit", as_amount(ws.cell(row, 3).value)),
            ("103", "FDR Profit", as_amount(ws.cell(row, 4).value)),
            ("104", "Sale", as_amount(ws.cell(row, 5).value)),
            ("105", "Service Charge", as_amount(ws.cell(row, 6).value)),
            ("105", "Rent", as_amount(ws.cell(row, 7).value)),
        ]
        for code, name, amount in vals:
            if amount <= 0:
                continue
            insert_voucher(
                cursor,
                voucher_type="RV",
                account_code=code,
                description=f"{name} - {month_label}",
                amount=amount,
                voucher_date=date_guess,
                notes="Imported monthly income from Year-2022(1)",
            )
            inserted += 1

    # Service charge account section
    sc_amount = as_amount(ws.cell(302, 10).value)
    if sc_amount > 0:
        payee_name = "Ms. Milu Begum"
        insert_voucher(
            cursor,
            voucher_type="RV",
            account_code="105",
            description=f"Service Charge Account - {payee_name}",
            amount=sc_amount,
            voucher_date="2022-06-30",
            notes="Imported service charge account section",
            payee_payor=payee_name,
        )
        inserted += 1

    return inserted


def run_import(excel_file: Path, apply_changes: bool) -> None:
    ensure_db_ready()

    if not excel_file.exists():
        raise FileNotFoundError(f"Workbook not found: {excel_file}")

    wb = openpyxl.load_workbook(excel_file, data_only=True)
    ws = wb["Year-2022(1)"]

    conn = get_connection()
    cursor = conn.cursor()

    counts = {
        "vouchers": 0,
        "investments": 0,
        "flatholders": 0,
        "flatholder_payments": 0,
        "source_rows": 0,
    }

    counts["source_rows"] += archive_workbook_rows(wb, cursor)

    try:
        counts["vouchers"] += import_expense_section(ws, cursor)
        inv_count, inv_vouchers = import_investments(ws, cursor)
        counts["investments"] += inv_count
        counts["vouchers"] += inv_vouchers

        holders, holder_payments = import_flatholders(ws, cursor)
        counts["flatholders"] += holders
        counts["flatholder_payments"] += holder_payments

        counts["vouchers"] += import_income_statement(ws, cursor)

        if apply_changes:
            backup = backup_db()
            conn.commit()
            print(f"Applied import successfully. Backup created: {backup}")
        else:
            conn.rollback()
            print("Dry run complete. No changes were written.")

        print("Import summary:")
        for key, value in counts.items():
            print(f"- {key}: {value}")

    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


def run_archive_only(excel_file: Path, apply_changes: bool) -> None:
    ensure_db_ready()

    if not excel_file.exists():
        raise FileNotFoundError(f"Workbook not found: {excel_file}")

    wb = openpyxl.load_workbook(excel_file, data_only=True)
    conn = get_connection()
    cursor = conn.cursor()
    try:
        archived = archive_workbook_rows(wb, cursor)
        if apply_changes:
            backup = backup_db()
            conn.commit()
            print(f"Archived workbook rows successfully. Backup created: {backup}")
        else:
            conn.rollback()
            print("Dry run complete. No changes were written.")
        print(f"- source_rows: {archived}")
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


def main() -> None:
    parser = argparse.ArgumentParser(description="Import APAN NIBASH Excel workbook into database")
    parser.add_argument("--file", required=True, help="Path to workbook (.xlsx)")
    mode = parser.add_mutually_exclusive_group(required=True)
    mode.add_argument("--apply", action="store_true", help="Write changes to database")
    mode.add_argument("--dry-run", action="store_true", help="Parse and simulate import only")
    mode.add_argument("--archive-only", action="store_true", help="Archive raw workbook rows only")
    args = parser.parse_args()

    if args.archive_only:
        run_archive_only(Path(args.file), apply_changes=True)
    else:
        run_import(Path(args.file), apply_changes=args.apply)


if __name__ == "__main__":
    main()
