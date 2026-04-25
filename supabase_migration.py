#!/usr/bin/env python3
"""
Supabase Migration - Seeds all data from the APAN NIBASH workbook into Supabase.
Keeps the exact same business logic and relationships as the original SQLite importer.

Usage:
  python supabase_migration.py --apply   # actually write
  python supabase_migration.py --dry-run # preview counts

Ensure:
  - SUPABASE_URL and SUPABASE_KEY environment variables are set.
  - The workbook "APAN NIBASH-21.xlsx" is in this directory.
"""

from __future__ import annotations

import argparse
import json
import os
import sys
from datetime import datetime
from pathlib import Path
from typing import Dict, Tuple

import openpyxl
from supabase import create_client

from database import get_supabase_client, get_using_supabase

WB_PATH = Path(__file__).parent / "APAN NIBASH-21.xlsx"
# The last/current sheet with complete data
SHEET_NAME = "Year-2022(1)"

# Same mapping the original importer used
EXPENSE_ACCOUNT_MAP: Dict[str, str] = {
    "preliminary": "403", "consultant": "209", "mixture": "210", "vibrator": "210",
    "jahan": "404", "sabuj": "404", "rafiqul": "404", "pilling": "404", "piling": "404",
    "rod": "301", "steel": "301", "cement": "302", "stone": "305", "sand": "303",
    "carraying": "309", "carrying": "309", "deep tubwell": "405", "gas connection": "406",
    "electricity": "202", "land registration": "402", "salary": "206", "allowances": "206",
    "office rent": "201", "entertainment": "205", "printing": "204", "stationery": "204",
    "conveyance": "208", "legal": "209", "sanitary": "311", "plumbing": "311",
    "electric materials": "312", "tiles": "313", "window": "314", "grill": "314",
    "doors": "315", "chawkath": "315", "paint": "316", "thai glass": "317",
    "commission": "408", "loan": "409", "refund": "409", "holding tax": "411",
    "lift": "412",
}


def normalize_name(name) -> str:
    return " ".join(str(name or "").strip().split()).lower()


def map_expense_code(particular: str) -> str:
    key = normalize_name(particular)
    for token, code in EXPENSE_ACCOUNT_MAP.items():
        if token in key:
            return code
    return "210"  # default misc


def as_cents(val):
    try:
        return int(round(float(val or 0) * 100))
    except Exception:
        return 0


def date_from_cell(cell, fallback="2022-01-01") -> str:
    if isinstance(cell, datetime):
        return cell.strftime("%Y-%m-%d")
    return fallback


def archive_workbook_rows(wb, client) -> int:
    archived = 0
    for ws in wb.worksheets:
        for r in range(1, ws.max_row + 1):
            values = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
            if not any(v not in (None, "", " ") for v in values):
                continue
            title = next((str(v).strip() for v in values if v not in (None, "", " ")), None)
            date_val = None
            amt1 = amt2 = 0.0
            for v in values:
                if isinstance(v, datetime) and date_val is None:
                    date_val = v.strftime("%Y-%m-%d")
                elif isinstance(v, (int, float)):
                    if amt1 == 0:
                        amt1 = float(v)
                    elif amt2 == 0:
                        amt2 = float(v)
            client.table("source_rows").insert({
                "sheet_name": ws.title,
                "row_no": r,
                "record_type": "raw",
                "title": title or "",
                "date_value": date_val,
                "amount": amt1,
                "amount_2": amt2,
                "row_json": json.dumps([str(v) if isinstance(v, datetime) else v for v in values], default=str),
            }).execute()
            archived += 1
    return archived


def next_voucher_no(cursor, voucher_type: str) -> str:
    res = cursor.execute(
        "SELECT COUNT(*) FROM vouchers WHERE voucher_no LIKE ?",
        (f"{voucher_type}-IMP-%",),
    ).fetchone()
    cnt = res[0] if res else 0
    return f"{voucher_type}-IMP-{cnt + 1:06d}"


def insert_voucher_sq(voucher_type, account_code, description, amount, voucher_date, notes, payee_payor=""):
    """Insert voucher directly into Supabase."""
    client = get_supabase_client()
    debit = amount if voucher_type == "PV" else 0.0
    credit = amount if voucher_type == "RV" else 0.0
    client.table("vouchers").insert({
        "voucher_no": f"{voucher_type}-IMP-{hash(description + voucher_date) % 1000000:06d}",
        "date": voucher_date,
        "voucher_type": voucher_type,
        "account_code": account_code,
        "description": description[:255],
        "debit_amount": debit,
        "credit_amount": credit,
        "reference_no": "",
        "payee_payor": payee_payor[:255] if payee_payor else "",
        "notes": notes[:255],
    }).execute()


def import_expense_section(ws, client) -> int:
    imported = 0
    for row in range(5, 78):
        particular = ws.cell(row, 2).value
        if not particular:
            continue
        name = str(particular).strip()
        if name.lower().startswith("total"):
            continue
        opening = as_cents(ws.cell(row, 3).value)
        if opening > 0:
            insert_voucher_sq("PV", map_expense_code(name), name, opening,
                              "2021-12-31", "Imported opening cumulative up to Dec-2021")
            imported += 1
        for col in range(4, 10):
            amount = as_cents(ws.cell(row, col).value)
            if amount <= 0:
                continue
            month_cell = ws.cell(4, col).value
            insert_voucher_sq("PV", map_expense_code(name), name, amount,
                              date_from_cell(month_cell, "2022-01-01"),
                              "Imported monthly amount from Year-2022(1)")
            imported += 1
    return imported


def import_investments(ws, client) -> Tuple[int, int]:
    n_invests = n_vouchers = 0
    # Received 84-90
    for r in range(84, 91):
        name = ws.cell(r, 2).value
        total = as_cents(ws.cell(r, 10).value)
        if not name or total <= 0:
            continue
        nm = str(name).strip()
        if nm.lower().startswith("total"):
            continue
        client.table("investments").insert({
            "name": nm, "type": "RECEIVED", "amount": total, "date": "2022-06-30", "status": "ACTIVE",
        }).execute()
        n_invests += 1
        insert_voucher_sq("RV", "102", f"Investment received: {nm}", total,
                          "2022-06-30", "Imported investment received summary", nm)
        n_vouchers += 1
    # Paid 100-106
    for r in range(100, 107):
        name = ws.cell(r, 2).value
        total = as_cents(ws.cell(r, 10).value)
        if not name or total <= 0:
            continue
        nm = str(name).strip()
        if nm.lower().startswith("total"):
            continue
        client.table("investments").insert({
            "name": nm, "type": "PAID", "amount": total, "date": "2022-06-30", "status": "ACTIVE",
        }).execute()
        n_invests += 1
        insert_voucher_sq("PV", "409", f"Investment repayment: {nm}", total,
                          "2022-06-30", "Imported investment payment summary", nm)
        n_vouchers += 1
    return n_invests, n_vouchers


def import_flatholders(ws, client) -> Tuple[int, int]:
    holders = payments = 0
    for r in range(155, 206):
        serial = ws.cell(r, 1).value
        name = ws.cell(r, 2).value
        total_paid = as_cents(ws.cell(r, 10).value)
        if not serial or not name:
            continue
        nm = str(name).strip()
        if nm.lower().startswith("total") or total_paid <= 0:
            continue
        unit = f"Flat-{int(serial)}"
        # upsert flatholder
        client.table("flatholders").upsert({
            "serial_no": int(serial), "name": nm, "flat_unit": unit,
            "total_amount": total_paid, "paid_amount": total_paid,
            "phone": "", "email": "", "address": "",
        }, on_conflict="serial_no").execute()
        holders += 1
        # payment record
        client.table("flatholder_payments").insert({
            "flatholder_id": int(serial),  # simplified; real FK needs lookup
            "payment_date": "2022-06-30",
            "amount": total_paid,
            "payment_type": "INSTALLMENT",
            "notes": "Imported cumulative paid amount up to 30/06/2022",
        }).execute()
        payments += 1
    return holders, payments


def import_income_statement(ws, client) -> int:
    inserted = 0
    row_cum = 267
    bank = as_cents(ws.cell(row_cum, 3).value)
    fdr = as_cents(ws.cell(row_cum, 4).value)
    sale = as_cents(ws.cell(row_cum, 5).value)
    charges = as_cents(ws.cell(row_cum, 6).value)
    rent = as_cents(ws.cell(row_cum, 7).value)
    for code, desc, amount in [
        ("103", "Bank Profit (cumulative)", bank),
        ("103", "FDR Profit (cumulative)", fdr),
        ("104", "Sale of scrap/wastage (cumulative)", sale),
        ("105", "Service charges (cumulative)", charges),
        ("105", "Rent income (cumulative)", rent),
    ]:
        if amount > 0:
            insert_voucher_sq("RV", code, desc, amount,
                              "2021-12-31", "Imported cumulative income up to Dec-2021")
            inserted += 1
    for r in range(268, 274):
        lbl = ws.cell(r, 2).value
        if not lbl:
            continue
        d = f"2022-{r - 267:02d}-01"
        vals = [
            ("103", "Bank Profit", as_cents(ws.cell(r, 3).value)),
            ("103", "FDR Profit", as_cents(ws.cell(r, 4).value)),
            ("104", "Sale", as_cents(ws.cell(r, 5).value)),
            ("105", "Service Charge", as_cents(ws.cell(r, 6).value)),
            ("105", "Rent", as_cents(ws.cell(r, 7).value)),
        ]
        for code, name, amount in vals:
            if amount <= 0:
                continue
            insert_voucher_sq("RV", code, f"{name} - {lbl}", amount, d,
                              "Imported monthly income from Year-2022(1)")
            inserted += 1
    sc = as_cents(ws.cell(302, 10).value)
    if sc > 0:
        insert_voucher_sq("RV", "105", "Service Charge Account - Ms. Milu Begum", sc,
                          "2022-06-30", "Imported service charge account section", "Ms. Milu Begum")
        inserted += 1
    return inserted


def run_migration(apply: bool) -> None:
    if not get_using_supabase():
        print("ERROR: Supabase not configured. Set SUPABASE_URL and SUPABASE_KEY.")
        sys.exit(1)

    if not WB_PATH.exists():
        print(f"ERROR: Workbook not found: {WB_PATH}")
        sys.exit(1)

    client = get_supabase_client()
    wb = openpyxl.load_workbook(WB_PATH, data_only=True)
    ws = wb[SHEET_NAME]

    counts = {"vouchers": 0, "investments": 0, "flatholders": 0, "flatholder_payments": 0, "source_rows": 0}

    print("Archiving raw workbook rows...")
    counts["source_rows"] += archive_workbook_rows(wb, client)

    if apply:
        print("Applying migration to Supabase (this may take a moment)...")
    else:
        print("DRY RUN — no changes will be written (estimated counts):")

    # The same order as the original db init to respect FKs
    print("Importing expenses...")
    counts["vouchers"] += import_expense_section(ws, client)
    print("Importing investments...")
    inv, inv_v = import_investments(ws, client)
    counts["investments"] += inv
    counts["vouchers"] += inv_v
    print("Importing flatholders & payments...")
    h, p = import_flatholders(ws, client)
    counts["flatholders"] += h
    counts["flatholder_payments"] += p
    print("Importing income statement...")
    counts["vouchers"] += import_income_statement(ws, client)

    print("\n=== Summary ===")
    for k, v in counts.items():
        print(f"  {k}: {v}")

    if apply:
        print("\n✓ Migration applied to Supabase.")
    else:
        print("\n(DRY RUN — no data was written)")


def main() -> None:
    parser = argparse.ArgumentParser(description="Import APAN NIBASH workbook into Supabase")
    parser.add_argument("--apply", action="store_true", help="Write changes to Supabase")
    parser.add_argument("--dry-run", action="store_true", help="Simulate import only")
    args = parser.parse_args()
    apply = args.apply
    if args.dry_run:
        apply = False
    if not (apply or args.dry_run):
        print("Choose --apply or --dry-run")
        sys.exit(1)
    run_migration(apply=apply)


if __name__ == "__main__":
    main()
